#!/usr/bin/env python3
"""
Unified Excel export module for all grant scrapers.
Provides consistent formatting across EU, Eureka, and other funding sources.
"""

import re
from datetime import datetime
from typing import List, Dict, Any, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    import sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


def clean_html(text: str) -> str:
    """Remove HTML tags and normalize whitespace."""
    if not text:
        return ""
    text = re.sub(r'<[^>]+>', ' ', text)
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


# Standard column definitions - order matters
STANDARD_COLUMNS = [
    {'key': 'source', 'header': 'Source', 'width': 15},
    {'key': 'title', 'header': 'Title', 'width': 50},
    {'key': 'status', 'header': 'Status', 'width': 12},
    {'key': 'identifier', 'header': 'Identifier', 'width': 25},
    {'key': 'programme', 'header': 'Programme', 'width': 20},
    {'key': 'budget', 'header': 'Budget', 'width': 15},
    {'key': 'open_date', 'header': 'Open Date', 'width': 22},
    {'key': 'close_date', 'header': 'Close Date', 'width': 22},
    {'key': 'all_deadlines', 'header': 'All Deadlines', 'width': 50},
    {'key': 'deadline_model', 'header': 'Deadline Model', 'width': 15},
    {'key': 'duration', 'header': 'Duration', 'width': 15},
    {'key': 'tags', 'header': 'Tags', 'width': 30},
    {'key': 'description', 'header': 'Description', 'width': 60},
    {'key': 'url', 'header': 'URL', 'width': 40},
]

# Status color mapping
STATUS_COLORS = {
    'Open': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # Green
    'Forthcoming': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),  # Yellow
    'Closed': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),  # Red
    'Active': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # Green
    'Upcoming': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),  # Yellow
}


def normalize_grant_for_excel(grant: Dict[str, Any], source_name: str) -> Dict[str, Any]:
    """
    Normalize any grant format to the standard Excel format.
    Works with EU grants, Eureka grants, or any other scraper output.

    Args:
        grant: Raw grant data (can be in various formats)
        source_name: Name of the source (e.g., 'horizon_europe', 'eureka')

    Returns:
        Normalized dictionary with standard column keys
    """
    # Handle different grant formats
    raw = grant.get('raw', {})
    metadata = raw.get('metadata', {}) if isinstance(raw, dict) else {}

    # Extract budget - try multiple formats
    budget = ""
    if 'budget' in metadata:
        budget_field = metadata['budget']
        if isinstance(budget_field, list) and budget_field:
            try:
                budget = f"€{int(budget_field[0]):,}"
            except (ValueError, TypeError):
                budget = str(budget_field[0])
    elif 'funding_info' in raw:
        budget = raw.get('funding_info', '')
    elif 'budget' in grant:
        budget = str(grant['budget'])

    # Extract description
    desc = ""
    if 'descriptionByte' in metadata:
        desc_field = metadata['descriptionByte']
        if isinstance(desc_field, list) and desc_field:
            desc = clean_html(desc_field[0])
        elif isinstance(desc_field, str):
            desc = clean_html(desc_field)
    elif 'description' in raw:
        desc = clean_html(raw.get('description', ''))
    elif 'description' in grant:
        desc = clean_html(grant.get('description', ''))

    # Truncate long descriptions
    if len(desc) > 1000:
        desc = desc[:1000] + "..."

    # Extract tags
    tags = metadata.get('crossCuttingPriorities', [])
    if not tags:
        tags = grant.get('tags', [])
    tags_str = ", ".join(tags) if isinstance(tags, list) else str(tags) if tags else ""

    # Format deadline_dates (multiple cutoff dates)
    deadline_dates = grant.get('deadline_dates', [])
    all_deadlines = ""
    if deadline_dates and isinstance(deadline_dates, list):
        all_deadlines = " | ".join(deadline_dates)

    # Extract identifier
    identifier = ""
    if 'identifier' in metadata:
        id_field = metadata['identifier']
        identifier = id_field[0] if isinstance(id_field, list) and id_field else str(id_field) if id_field else ""
    elif 'id' in grant:
        identifier = str(grant['id'])

    # Extract programme
    programme = ""
    if 'programme' in metadata:
        programme = str(metadata['programme'])
    elif 'programme' in grant:
        programme = str(grant['programme']) if grant.get('programme') else ""

    # Extract deadline model
    deadline_model = ""
    if 'deadlineModel' in metadata:
        dm_field = metadata['deadlineModel']
        deadline_model = dm_field[0] if isinstance(dm_field, list) and dm_field else str(dm_field) if dm_field else ""

    # Extract duration
    duration = ""
    if 'duration' in metadata:
        dur_field = metadata['duration']
        duration = clean_html(dur_field[0]) if isinstance(dur_field, list) and dur_field else str(dur_field) if dur_field else ""

    # Map status to readable format
    status = grant.get('status', '')
    status_map = {
        '31094501': 'Forthcoming',
        '31094502': 'Open',
        '31094503': 'Closed',
    }
    if isinstance(status, str):
        match = re.search(r"'(\d+)'", status)
        if match:
            status = status_map.get(match.group(1), status)

    return {
        'source': source_name.replace('_', ' ').title(),
        'title': grant.get('title', ''),
        'status': status,
        'identifier': identifier,
        'programme': programme,
        'budget': budget,
        'open_date': grant.get('open_date', ''),
        'close_date': grant.get('close_date', ''),
        'all_deadlines': all_deadlines,
        'deadline_model': deadline_model,
        'duration': duration,
        'tags': tags_str,
        'description': desc,
        'url': grant.get('url', ''),
    }


def export_grants_to_excel(
    grants: List[Dict[str, Any]],
    output_path: str,
    sheet_name: str = "Grants",
    columns: Optional[List[Dict]] = None
) -> str:
    """
    Export grants to a formatted Excel file.

    Args:
        grants: List of normalized grant dictionaries
        output_path: Path for the output Excel file
        sheet_name: Name for the worksheet
        columns: Optional custom column definitions (uses STANDARD_COLUMNS if not provided)

    Returns:
        Path to the saved file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    cols = columns or STANDARD_COLUMNS
    headers = [c['header'] for c in cols]
    keys = [c['key'] for c in cols]

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Write data
    for row_idx, grant in enumerate(grants, 2):
        for col_idx, key in enumerate(keys, 1):
            value = grant.get(key, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=True)

            # Color status cells
            if key == 'status' and value in STATUS_COLORS:
                cell.fill = STATUS_COLORS[value]

    # Set column widths
    for col_idx, col_def in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_def['width']

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    # Save
    wb.save(output_path)
    return output_path


def generate_summary(grants: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Generate summary statistics for the scraped grants."""
    summary = {
        'total': len(grants),
        'by_status': {},
        'by_source': {},
        'with_multiple_deadlines': 0,
    }

    for g in grants:
        # Count by status
        status = g.get('status', 'Unknown')
        summary['by_status'][status] = summary['by_status'].get(status, 0) + 1

        # Count by source
        source = g.get('source', 'Unknown')
        summary['by_source'][source] = summary['by_source'].get(source, 0) + 1

        # Count multiple deadlines
        if g.get('all_deadlines'):
            summary['with_multiple_deadlines'] += 1

    return summary


def print_summary(summary: Dict[str, Any]):
    """Print a formatted summary to console."""
    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    print(f"Total grants: {summary['total']}")

    if summary['by_source']:
        print("\nBy source:")
        for source, count in sorted(summary['by_source'].items()):
            print(f"  - {source}: {count}")

    if summary['by_status']:
        print("\nBy status:")
        for status, count in sorted(summary['by_status'].items()):
            print(f"  - {status}: {count}")

    if summary['with_multiple_deadlines']:
        print(f"\nGrants with multiple deadlines: {summary['with_multiple_deadlines']}")
